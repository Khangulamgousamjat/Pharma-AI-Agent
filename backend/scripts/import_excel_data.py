import sys
import os
from datetime import datetime, date, timedelta
import openpyxl

# Add parent dir to path to import app modules
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.database import SessionLocal, init_db
from app.models.medicine import Medicine

def parse_date(date_val):
    if not date_val:
        return date.today() + timedelta(days=365) # Default 1 year expiry
    if isinstance(date_val, datetime):
        return date_val.date()
    if isinstance(date_val, date):
        return date_val
    try:
        # Try parsing 'YYYY-MM-DD'
        return datetime.strptime(str(date_val).strip(), "%Y-%m-%d").date()
    except Exception:
        return date.today() + timedelta(days=365)

def main():
    print("Running Medicine Database Importer...")
    init_db()
    db = SessionLocal()
    
    # 1. Build a prescription map from "Consumer Order History 1.xlsx"
    rx_map = {}
    history_path = "../Consumer Order History 1.xlsx"
    if os.path.exists(history_path):
        try:
            wb = openpyxl.load_workbook(history_path, data_only=True)
            sheet = wb.active
            # Row 5 is header, data starts from Row 6
            for row_idx in range(6, sheet.max_row + 1):
                product_name = sheet.cell(row=row_idx, column=5).value
                rx_val = sheet.cell(row=row_idx, column=9).value
                if product_name and rx_val:
                    is_rx = str(rx_val).strip().lower() in ("yes", "true", "1")
                    rx_map[str(product_name).strip()] = is_rx
            print(f"Loaded {len(rx_map)} prescription mappings from order history.")
        except Exception as e:
            print(f"Could not load order history prescription map: {e}")
            
    # List of keywords for prescription requirement (Rx)
    rx_keywords = [
        "antibiotic", "infection", "blood pressure", "hypertension", "diabetes", 
        "metformin", "amoxicillin", "ciprofloxacin", "prescription", "rx", "cholesterol",
        "atorvastatin", "insulin", "antidepressant", "asthma", "steroid"
    ]
    
    def check_rx(name, desc):
        name_lower = name.lower()
        desc_lower = (desc or "").lower()
        
        # Check order history map first
        if name in rx_map:
            return rx_map[name]
        
        # Check keywords
        for kw in rx_keywords:
            if kw in name_lower or kw in desc_lower:
                return True
        return False

    imported_count = 0
    skipped_count = 0

    # 2. Import "200_medicines_database.xlsx"
    db_200_path = "../200_medicines_database.xlsx"
    if os.path.exists(db_200_path):
        print(f"Processing {db_200_path}...")
        try:
            wb = openpyxl.load_workbook(db_200_path, data_only=True)
            sheet = wb.active
            seen_in_session = set()
            # Headers: product id, product name, price rec, package size, descriptions, exp date
            for row_idx in range(2, sheet.max_row + 1):
                name = sheet.cell(row=row_idx, column=2).value
                price = sheet.cell(row=row_idx, column=3).value
                unit = sheet.cell(row=row_idx, column=4).value
                desc = sheet.cell(row=row_idx, column=5).value
                exp_date_val = sheet.cell(row=row_idx, column=6).value
                
                if not name:
                    continue
                
                name = str(name).strip()
                if name in seen_in_session:
                    skipped_count += 1
                    continue
                
                # Check if exists in DB
                existing = db.query(Medicine).filter(Medicine.name == name).first()
                if existing:
                    skipped_count += 1
                    continue
                
                seen_in_session.add(name)
                
                try:
                    price_val = float(price) if price is not None else 10.0
                except ValueError:
                    price_val = 10.0
                    
                is_rx = check_rx(name, desc)
                exp_date = parse_date(exp_date_val)
                
                med = Medicine(
                    name=name,
                    price=price_val,
                    unit=str(unit or "tablets").strip(),
                    description=str(desc or "").strip(),
                    expiry_date=exp_date,
                    stock=100, # Default stock
                    prescription_required=is_rx
                )
                db.add(med)
                imported_count += 1
            db.commit()
            print("Successfully processed 200 medicines database.")
        except Exception as e:
            print(f"Error importing 200 medicines: {e}")
            db.rollback()

    # 3. Import "products-export.xlsx"
    prod_export_path = "../products-export.xlsx"
    if os.path.exists(prod_export_path):
        print(f"Processing {prod_export_path}...")
        try:
            wb = openpyxl.load_workbook(prod_export_path, data_only=True)
            sheet = wb.active
            seen_in_session = set()
            # Headers: product id, product name, pzn, price rec, package size, descriptions
            for row_idx in range(2, sheet.max_row + 1):
                name = sheet.cell(row=row_idx, column=2).value
                price = sheet.cell(row=row_idx, column=4).value
                unit = sheet.cell(row=row_idx, column=5).value
                desc = sheet.cell(row=row_idx, column=6).value
                
                if not name:
                    continue
                
                name = str(name).strip()
                if name in seen_in_session:
                    skipped_count += 1
                    continue
                
                # Check if exists in DB
                existing = db.query(Medicine).filter(Medicine.name == name).first()
                if existing:
                    skipped_count += 1
                    continue
                
                seen_in_session.add(name)
                
                try:
                    price_val = float(price) if price is not None else 15.0
                except ValueError:
                    price_val = 15.0
                    
                is_rx = check_rx(name, desc)
                exp_date = date.today() + timedelta(days=365 * 2) # Default 2 year expiry
                
                med = Medicine(
                    name=name,
                    price=price_val,
                    unit=str(unit or "units").strip(),
                    description=str(desc or "").strip(),
                    expiry_date=exp_date,
                    stock=100, # Default stock
                    prescription_required=is_rx
                )
                db.add(med)
                imported_count += 1
            db.commit()
            print("Successfully processed products export.")
        except Exception as e:
            print(f"Error importing products export: {e}")
            db.rollback()

    db.close()
    print(f"Done. Imported: {imported_count}, Skipped (duplicate names): {skipped_count}")

if __name__ == "__main__":
    main()

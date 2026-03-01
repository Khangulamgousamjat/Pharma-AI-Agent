import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from app.services.order_service import get_all_orders, get_user_orders
from app.services.medicine_service import get_all_medicines
from tempfile import NamedTemporaryFile
import os


def _style_header_row(ws, num_cols: int):
    fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center")
    for col_num in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _auto_fit_columns(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def _save_wb(wb) -> str:
    temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    return temp_file.name


def export_orders_to_excel(db: Session) -> str:
    """Exports all orders (admin) to an Excel file."""
    orders = get_all_orders(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Orders"

    headers = ["Order ID", "User ID", "Medicine ID", "Quantity", "Total Price (₹)", "Status", "Date"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for order in orders:
        ws.append([
            order.id,
            order.user_id,
            order.medicine_id,
            order.quantity,
            order.total_price,
            order.status,
            order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else ""
        ])

    _auto_fit_columns(ws)
    return _save_wb(wb)


def export_user_orders_to_excel(db: Session, user_id: int) -> str:
    """Exports orders for a specific user to an Excel file."""
    orders = get_user_orders(db, user_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Orders"

    headers = ["Order ID", "Medicine", "Quantity", "Total Price (₹)", "Status", "Date"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for order in orders:
        med_name = order.medicine.name if order.medicine else f"Medicine #{order.medicine_id}"
        ws.append([
            order.id,
            med_name,
            order.quantity,
            order.total_price,
            order.status,
            order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else ""
        ])

    _auto_fit_columns(ws)
    return _save_wb(wb)


def export_inventory_to_excel(db: Session) -> str:
    """Exports the full medicine inventory to an Excel file."""
    medicines = get_all_medicines(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = ["ID", "Name", "Stock", "Unit", "Price (₹)", "Type", "Expiry Date", "Description"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for med in medicines:
        ws.append([
            med.id,
            med.name,
            med.stock,
            med.unit,
            med.price,
            "Rx" if med.prescription_required else "OTC",
            str(med.expiry_date) if med.expiry_date else "",
            med.description or ""
        ])

    _auto_fit_columns(ws)
    return _save_wb(wb)

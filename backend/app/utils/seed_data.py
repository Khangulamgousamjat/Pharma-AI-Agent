"""
utils/seed_data.py — Seed initial medicine inventory into the database.

Run on application startup to ensure the database has sample medicines
for testing and demonstration. Uses upsert logic (skip if already exists).
"""

from sqlalchemy.orm import Session
from datetime import date
import logging

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Sample Medicine Catalogue
# ---------------------------------------------------------------------------
# Medicines are split into:
#   - OTC (Over The Counter): prescription_required = False → agent allows order
#   - Rx (Prescription):      prescription_required = True  → agent rejects order
# ---------------------------------------------------------------------------
SEED_MEDICINES = [
    # OTC Medicines — agent can order these freely
    {
        "name": "Paracetamol 500mg",
        "stock": 500,
        "unit": "tablets",
        "price": 5.0,
        "prescription_required": False,
        "expiry_date": date(2026, 12, 31),
        "description": "Fever and pain relief. OTC medicine.",
    },
    {
        "name": "Ibuprofen 400mg",
        "stock": 300,
        "unit": "tablets",
        "price": 8.0,
        "prescription_required": False,
        "expiry_date": date(2026, 10, 31),
        "description": "Anti-inflammatory and pain relief. OTC.",
    },
    {
        "name": "Cetirizine 10mg",
        "stock": 200,
        "unit": "tablets",
        "price": 6.0,
        "prescription_required": False,
        "expiry_date": date(2026, 8, 31),
        "description": "Antihistamine for allergies. OTC.",
    },
    {
        "name": "Cough Syrup 100ml",
        "stock": 150,
        "unit": "bottles",
        "price": 45.0,
        "prescription_required": False,
        "expiry_date": date(2026, 6, 30),
        "description": "Relieves dry and productive coughs. OTC.",
    },
    {
        "name": "Vitamin C 1000mg",
        "stock": 400,
        "unit": "tablets",
        "price": 12.0,
        "prescription_required": False,
        "expiry_date": date(2027, 1, 31),
        "description": "Immunity booster. OTC supplement.",
    },
    {
        "name": "Antacid Tablets",
        "stock": 250,
        "unit": "tablets",
        "price": 4.0,
        "prescription_required": False,
        "expiry_date": date(2026, 9, 30),
        "description": "Relieves heartburn and indigestion. OTC.",
    },
    {
        "name": "ORS Sachets",
        "stock": 100,
        "unit": "sachets",
        "price": 10.0,
        "prescription_required": False,
        "expiry_date": date(2026, 12, 31),
        "description": "Oral rehydration salts for dehydration. OTC.",
    },
    # Rx Medicines — agent will reject without prescription
    {
        "name": "Amoxicillin 500mg",
        "stock": 200,
        "unit": "capsules",
        "price": 25.0,
        "prescription_required": True,
        "expiry_date": date(2026, 5, 31),
        "description": "Antibiotic. Prescription required.",
    },
    {
        "name": "Metformin 500mg",
        "stock": 180,
        "unit": "tablets",
        "price": 15.0,
        "prescription_required": True,
        "expiry_date": date(2026, 11, 30),
        "description": "Diabetes management. Prescription required.",
    },
    {
        "name": "Atorvastatin 20mg",
        "stock": 120,
        "unit": "tablets",
        "price": 35.0,
        "prescription_required": True,
        "expiry_date": date(2026, 7, 31),
        "description": "Cholesterol management. Prescription required.",
    },
]


def seed_medicines(db: Session) -> None:
    """
    Seed the medicines table with sample data and excel data.

    Skips medicines that already exist (by name) to prevent duplicates.
    Safe to call on every startup.

    Args:
        db: SQLAlchemy database session
    """
    from app.models.medicine import Medicine
    from datetime import date, timedelta
    import os
    import openpyxl

    # 1. Seed hardcoded SEED_MEDICINES
    seeded_count = 0
    for med_data in SEED_MEDICINES:
        # Check if medicine already exists by name
        existing = db.query(Medicine).filter(Medicine.name == med_data["name"]).first()
        if not existing:
            medicine = Medicine(**med_data)
            db.add(medicine)
            seeded_count += 1

    if seeded_count > 0:
        db.commit()
        logger.info(f"Seeded {seeded_count} hardcoded medicines into database.")
    else:
        logger.info("Medicine seed data already present, skipping.")

    # 2. Try to seed from Excel files if present
    history_paths = ["../Consumer Order History 1.xlsx", "./Consumer Order History 1.xlsx", "Consumer Order History 1.xlsx"]
    history_path = next((p for p in history_paths if os.path.exists(p)), None)
    
    rx_map = {}
    if history_path:
        try:
            wb = openpyxl.load_workbook(history_path, data_only=True)
            sheet = wb.active
            for row_idx in range(6, sheet.max_row + 1):
                product_name = sheet.cell(row=row_idx, column=5).value
                rx_val = sheet.cell(row=row_idx, column=9).value
                if product_name and rx_val:
                    is_rx = str(rx_val).strip().lower() in ("yes", "true", "1")
                    rx_map[str(product_name).strip()] = is_rx
            logger.info(f"Loaded {len(rx_map)} Rx mappings from order history.")
        except Exception as e:
            logger.error(f"Error loading order history Rx map: {e}")

    rx_keywords = [
        "antibiotic", "infection", "blood pressure", "hypertension", "diabetes", 
        "metformin", "amoxicillin", "ciprofloxacin", "prescription", "rx", "cholesterol",
        "atorvastatin", "insulin", "antidepressant", "asthma", "steroid"
    ]
    
    def check_rx(name, desc):
        name_lower = name.lower()
        desc_lower = (desc or "").lower()
        if name in rx_map:
            return rx_map[name]
        for kw in rx_keywords:
            if kw in name_lower or kw in desc_lower:
                return True
        return False

    def parse_date(date_val):
        if not date_val:
            return date.today() + timedelta(days=365)
        if isinstance(date_val, datetime):
            return date_val.date()
        if isinstance(date_val, date):
            return date_val
        try:
            from datetime import datetime as dt
            return dt.strptime(str(date_val).strip(), "%Y-%m-%d").date()
        except Exception:
            return date.today() + timedelta(days=365)

    excel_imported = 0
    excel_skipped = 0
    seen_in_session = set()

    # Import 200 medicines database
    db_200_paths = ["../200_medicines_database.xlsx", "./200_medicines_database.xlsx", "200_medicines_database.xlsx"]
    db_200_path = next((p for p in db_200_paths if os.path.exists(p)), None)
    
    if db_200_path:
        try:
            wb = openpyxl.load_workbook(db_200_path, data_only=True)
            sheet = wb.active
            for row_idx in range(2, sheet.max_row + 1):
                name = sheet.cell(row=row_idx, column=2).value
                price = sheet.cell(row=row_idx, column=3).value
                unit = sheet.cell(row=row_idx, column=4).value
                desc = sheet.cell(row=row_idx, column=5).value
                exp_date_val = sheet.cell(row=row_idx, column=6).value
                
                if not name:
                    continue
                name = str(name).strip()
                if name in seen_in_session:
                    excel_skipped += 1
                    continue
                
                existing = db.query(Medicine).filter(Medicine.name == name).first()
                if existing:
                    excel_skipped += 1
                    continue
                
                seen_in_session.add(name)
                
                try:
                    price_val = float(price) if price is not None else 10.0
                except ValueError:
                    price_val = 10.0
                    
                is_rx = check_rx(name, desc)
                exp_date = parse_date(exp_date_val)
                
                med = Medicine(
                    name=name,
                    price=price_val,
                    unit=str(unit or "tablets").strip(),
                    description=str(desc or "").strip(),
                    expiry_date=exp_date,
                    stock=100,
                    prescription_required=is_rx
                )
                db.add(med)
                excel_imported += 1
            db.commit()
            logger.info("Successfully seeded medicines from 200_medicines_database.xlsx")
        except Exception as e:
            logger.error(f"Error seeding from 200 medicines db: {e}")
            db.rollback()

    # Import products export
    prod_paths = ["../products-export.xlsx", "./products-export.xlsx", "products-export.xlsx"]
    prod_path = next((p for p in prod_paths if os.path.exists(p)), None)
    
    if prod_path:
        try:
            wb = openpyxl.load_workbook(prod_path, data_only=True)
            sheet = wb.active
            for row_idx in range(2, sheet.max_row + 1):
                name = sheet.cell(row=row_idx, column=2).value
                price = sheet.cell(row=row_idx, column=4).value
                unit = sheet.cell(row=row_idx, column=5).value
                desc = sheet.cell(row=row_idx, column=6).value
                
                if not name:
                    continue
                name = str(name).strip()
                if name in seen_in_session:
                    excel_skipped += 1
                    continue
                
                existing = db.query(Medicine).filter(Medicine.name == name).first()
                if existing:
                    excel_skipped += 1
                    continue
                
                seen_in_session.add(name)
                
                try:
                    price_val = float(price) if price is not None else 15.0
                except ValueError:
                    price_val = 15.0
                    
                is_rx = check_rx(name, desc)
                exp_date = date.today() + timedelta(days=365 * 2)
                
                med = Medicine(
                    name=name,
                    price=price_val,
                    unit=str(unit or "units").strip(),
                    description=str(desc or "").strip(),
                    expiry_date=exp_date,
                    stock=100,
                    prescription_required=is_rx
                )
                db.add(med)
                excel_imported += 1
            db.commit()
            logger.info("Successfully seeded medicines from products-export.xlsx")
        except Exception as e:
            logger.error(f"Error seeding from products export: {e}")
            db.rollback()

    logger.info(f"Excel seed summary: Imported={excel_imported}, Skipped={excel_skipped}")


def seed_admin_user(db: Session) -> None:
    """
    Create or update the default admin user.
    Admin credentials: admin@gmail.com / Kingkhan@12
    """
    from app.models.user import User
    from app.utils.security import hash_password

    admin_email = "admin@gmail.com"
    admin_password = "Kingkhan@12"

    admin = db.query(User).filter(User.email == admin_email).first()
    if not admin:
        admin = User(
            name="Admin",
            email=admin_email,
            password_hash=hash_password(admin_password),
            role="admin",
            is_approved=1
        )
        db.add(admin)
        db.commit()
        logger.info(f"Created admin user: {admin_email}")
    else:
        # Update existing admin with new credentials
        admin.password_hash = hash_password(admin_password)
        admin.role = "admin"
        admin.is_approved = 1
        db.commit()
        logger.info(f"Updated existing admin user: {admin_email}")


def seed_pharmacist_user(db: Session) -> None:
    """
    Phase 2: Create a default pharmacist user if none exists.

    Pharmacist credentials:
        Email: pharmacist@pharmaagent.com
        Password: pharma123

    The pharmacist can:
      - View pending prescriptions (GET /pharmacist/prescriptions/pending)
      - Approve prescriptions (POST /pharmacist/prescriptions/{id}/verify)

    Args:
        db: SQLAlchemy database session
    """
    from app.models.user import User
    from app.utils.security import hash_password

    existing = db.query(User).filter(User.email == "pharmacist@pharmaagent.com").first()
    if not existing:
        pharmacist = User(
            name="Dr. Pharmacist",
            email="pharmacist@pharmaagent.com",
            password_hash=hash_password("pharma123"),
            role="pharmacist",
            is_approved=1
        )
        db.add(pharmacist)
        db.commit()
        logger.info("Default pharmacist user created: pharmacist@pharmaagent.com / pharma123")
    else:
        logger.info("Pharmacist user already exists, skipping.")


def seed_demo_user(db: Session) -> None:
    """
    Create a default demo user (regular user role) if none exists.

    Demo credentials:
        Email: john@example.com
        Password: user123

    Used for the "user" role tab on the login page.
    """
    from app.models.user import User
    from app.utils.security import hash_password

    existing = db.query(User).filter(User.email == "john@example.com").first()
    if not existing:
        demo_user = User(
            name="John Demo",
            email="john@example.com",
            password_hash=hash_password("user123"),
            role="user",
            is_approved=1
        )
        db.add(demo_user)
        db.commit()
        logger.info("Demo user created: john@example.com / user123")
    else:
        logger.info("Demo user already exists, skipping.")

